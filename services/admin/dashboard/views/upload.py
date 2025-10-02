from __future__ import annotations

import csv
import io
from typing import List

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.utils.decorators import method_decorator
from openpyxl import load_workbook
from django_ratelimit.decorators import ratelimit
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from shared.core.forms import BulkUploadForm, NumberForm
from shared.core.models import Number

from django.conf import settings


@login_required
def upload_view(request: HttpRequest) -> HttpResponse:
    form = BulkUploadForm(request.POST or None, request.FILES or None)
    results = None
    if request.method == "POST" and form.is_valid():
        try:
            results = handle_upload(form.cleaned_data["file"], form.cleaned_data["dry_run"], form.cleaned_data["upsert"])
            if form.cleaned_data["dry_run"]:
                messages.info(request, f"Dry-run complete: {results['inserted']} insertable, {results['updated']} updatable, {results['errors']} errors.")
            else:
                messages.success(request, f"Upload complete: inserted {results['inserted']}, updated {results['updated']}, errors {results['errors']}.")
        except ValueError as exc:
            messages.error(request, str(exc))
    return render(request, "upload.html", {"form": form, "results": results})


def _read_rows(uploaded_file) -> List[dict]:
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if name.endswith(".csv"):
        text_file = io.TextIOWrapper(uploaded_file.file, encoding="utf-8")
        reader = csv.DictReader(text_file)
        return list(reader)
    if name.endswith(".xlsx"):
        workbook = load_workbook(uploaded_file, read_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.rows)]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows
    raise ValueError("Unsupported file format. Use CSV or XLSX.")


def handle_upload(uploaded_file, dry_run: bool, upsert: bool) -> dict:
    rows = _read_rows(uploaded_file)
    inserted = updated = errors = 0
    error_rows = []
    with transaction.atomic():
        for row in rows:
            data = {
                "area_code": str(row.get("area_code", "")).strip(),
                "phone_number": str(row.get("phone_number", "")).strip(),
                "cost": row.get("cost", 0),
            }
            form = NumberForm(data)
            if not form.is_valid():
                errors += 1
                error_rows.append({"row": row, "errors": form.errors})
                continue
            area_code = form.cleaned_data["area_code"]
            phone_number = form.cleaned_data["phone_number"]
            cost = form.cleaned_data["cost"]
            try:
                existing = Number.objects.get(area_code=area_code, phone_number=phone_number)
            except Number.DoesNotExist:
                if not dry_run:
                    form.save()
                inserted += 1
            else:
                if upsert:
                    if not dry_run:
                        existing.cost = cost
                        existing.save(update_fields=["cost", "updated_at"])
                    updated += 1
                else:
                    errors += 1
                    error_rows.append({"row": row, "errors": {"non_field_errors": ["Duplicate number"]}})
        if dry_run:
            transaction.set_rollback(True)
    return {"inserted": inserted, "updated": updated, "errors": errors, "error_rows": error_rows}


class BulkUploadApiView(APIView):
    permission_classes = [IsAuthenticated]

    @method_decorator(ratelimit(key="ip", rate=settings.ADMIN_API_RATE_LIMIT, method="POST", block=True))
    def post(self, request: HttpRequest) -> Response:
        form = BulkUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return Response({"error": form.errors}, status=400)
        try:
            results = handle_upload(form.cleaned_data["file"], form.cleaned_data["dry_run"], form.cleaned_data["upsert"])
        except ValueError as exc:
            return Response({"error": str(exc)}, status=400)
        return Response(results)
